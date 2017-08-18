# -*- coding:UTF-8 -*-
import sys
reload(sys)
sys.setdefaultencoding('utf8')
import matplotlib.pyplot as plt
import pandas as pd
import numpy as np
from openpyxl import Workbook
from matplotlib.ticker import MultipleLocator, FormatStrFormatter

"""
@function:将匹配合并结果和检测结果比较分类confidence变化
@param param1: 匹配合并列表文件
@param param2: 检测结果列表文件
@param param3: 待保存confidence变化文件
"""
def save_data(priorList, resultList, data_xlsx):
    with open(priorList) as fp1, open(resultList) as fp2: # 对于每个测试图片
        for resultFile in fp2: # 每一行匹配数据 resultFile
            priorFile = fp1.readline() # 每一行检测数据 priorFile
            prior_datas = priorFile.strip().split('\t')
            result_datas = resultFile.strip().split('\t')

            prior_IOU = []  # 训练时所有匹配的prior box与gt box的IOU
            prior_boxes_total = int(prior_datas[2]) # 匹配box数量
            for i in range(0, prior_boxes_total, 1):
                prior_IOU.append(float(prior_datas[7*i + 4]))

            # reslut_conf = []  # 检测得到result box置信度
            result_boxes_total = (len(result_datas)-1)/6  # 检测得到的box数量
            for k in range(0, len(conf_thresholds), 1):  # 对于每个分类置信阈值
                for j in range(0, result_boxes_total, 1): # 对于每个prior box
                    # index = int(prior_IOU[j] / thresholds[0])
                    for i in range(0, len(thresholds), 1): # 判断IOU区间段
                        if (prior_IOU[j] < thresholds[i]):
                            index = i
                            break
                    conf = float(result_datas[6 * j + 2]) #分类置信度
                    if conf >= conf_thresholds[k]:
                        all_change_group[k][index]['Pos'] += 1
                    else:
                        all_change_group[k][index]['Neg']  += 1

        # 将all_change_group数据写入xlsx表格中
        # 在内存中创建一个workbook对象，而且会至少创建一个 worksheet
        wb = Workbook()
        for k in range(0, len(conf_thresholds), 1):
            # 获取当前活跃的worksheet,默认就是第一个worksheet
            ws = wb.create_sheet()
            # 设置单元格的值，A1等于6(测试可知openpyxl的行和列编号从1开始计算)，B1等于7
            ws.cell(row=1, column=1).value = 'IOU'
            ws.cell("B1").value = 'Status'
            ws.cell("C1").value = 'NUM'
            # 从第1列第2行开始，写入IOU区间值
            for row in range(1, len(thresholds) + 1):
                ws.cell(row=row * 2, column=1).value = (thresholds[row - 1])
                ws.cell(row=row * 2 + 1, column=1).value = (thresholds[row - 1])
            # 从第2列第2行开始，写入Status值
            for row in range(1, len(thresholds) + 1):
                ws.cell(row=row * 2, column=2).value = 'Neg'
                ws.cell(row=row * 2 + 1, column=2).value = "Pos"
            # 从第3列第2行开始，写入NUM数据
            for row in range(1, len(thresholds) + 1):
                ws.cell(row=row * 2, column=3).value = all_change_group[k][row - 1]['Neg']
                ws.cell(row=row * 2 + 1, column=3).value = all_change_group[k][row - 1]['Pos']
        wb.save(filename=data_xlsx)


"""
@function:将各个conf阈值条件下confidence的变化以recall曲线展现
@param param1: confidence变化文件
"""
def show_bar(data_mat):
    fig1, ax1 = plt.subplots(1)
    prior_num = []
    for k in range(0, len(conf_thresholds), 1):
        recalls = []
        sheet_name = 'Sheet'+ str(k+1)
        df = pd.read_excel(data_mat, sheet_name)
        var = list(df['NUM'])
        for j in range(0, len(thresholds), 1):
            if k==0:
                prior_num.append(var[2*j+1] + var[2*j])
            recall = float(var[2*j+1]) / (float(var[2*j+1]) + float(var[2*j]))
            recalls.append(recall)
        ax1.plot(thresholds, recalls, lw=2, color=colors[k], label=str(thresholds[k])) # 绘制每一条recall曲线
        # plt.annotate(thresholds[k], xy=(thresholds[len(thresholds)/2], recalls[len(recalls)/2]),
        #              xytext=(thresholds[len(thresholds)/2], recalls[len(recalls)/2]),
        #              arrowprops = dict(facecolor="r", headlength=5, headwidth=5, width=2))

    plt.grid()
    plt.xlabel('IOU')
    plt.ylabel('Recall')
    #plt.title('IOU-Recall')
    plt.legend(loc="upper left")
    ax2 = ax1.twiny()
    ax2.set_xlim(0.1, 1)
    plt.xticks(thresholds, prior_num, rotation=10)
    plt.show()

#matplotlib.rcParams['figure.figsize'] = (6, 8)  # 设定显示大小
plt.rcParams['image.interpolation'] = 'nearest'
plt.rcParams['image.cmap'] = 'gray'
colors = plt.cm.hsv(np.linspace(0, 1, 10)).tolist()
ROOTDIR = "\\\\192.168.1.186/PedestrianData/" # 样本根目录
min_threshold = 0.1 # 最小IOU区间
thresholds = np.linspace( min_threshold, 1, 10 ) # IOU 区间
min_conf_threshold = 0.1 # 最小分类置信度阈值
conf_thresholds = np.linspace( min_conf_threshold, 1, 10 ) # 分类置信度阈值
all_change_group =  [[] for x in range(len(conf_thresholds))]  # 初始化
for j in range(0, len(conf_thresholds), 1):
    for i in range(0, len(thresholds), 1):
        all_change_group[j].append({'Neg': 0, 'Pos': 0})


if __name__ == "__main__":
    save_data("../Data_0810/IOU_ALL_image_List.txt", "../Data_0810/result_ALL_image_List.txt", "../Data_0810/confidence_change_statistic.xlsx")
    show_bar("../Data_0810/confidence_change_statistic.xlsx")