# -*- coding:UTF-8 -*-
import sys
reload(sys)
sys.setdefaultencoding('utf8')
import matplotlib.pyplot as plt
import pandas as pd
import numpy as np
from openpyxl import Workbook

"""
@function:将匹配合并结果和检测结果比较分类confidence变化
@param param1: 匹配合并列表文件
@param param2: 检测结果列表文件
@param param3: 待保存confidence变化文件
"""
def save_data(priorList, resultList, data_xlsx):
    with open(priorList) as fp1, open(resultList) as fp2: # 对于每个测试图片
        plt.close('all')
        for resultFile in fp2: # 每一行匹配数据 resultFile
            priorFile = fp1.readline() # 每一行检测数据 priorFile
            prior_datas = priorFile.strip().split('\t')
            result_datas = resultFile.strip().split('\t')

            prior_IOU = []  # 训练时所有匹配的prior box与gt box的IOU
            prior_boxes_total = int(prior_datas[2]) # 匹配box数量
            for i in range(0, prior_boxes_total, 1):
                prior_IOU.append(float(prior_datas[7*i + 4]))

            # reslut_conf = []  # 检测得到result box置信度
            result_boxes_total = (len(result_datas)-1)/6  # 检测得到的box数量
            for j in range(0, result_boxes_total, 1):
                #index = int(prior_IOU[j] / thresholds[0])
                for i in range(0, len(thresholds), 1):  # 判断IOU区间段
                    if (prior_IOU[j] < thresholds[i]):
                        index = i
                        break
                conf = float(result_datas[6 * j + 2]) #分类置信度
                # reslut_conf.append(conf)
                if conf >= conf_threshold:
                    all_change_group[index]['Pos'] += 1
                else:
                    all_change_group[index]['Neg']  += 1

    # 将all_change_group数据写入xlsx表格中
    # 在内存中创建一个workbook对象，而且会至少创建一个 worksheet
    wb = Workbook()
    # 获取当前活跃的worksheet,默认就是第一个worksheet
    ws = wb.active
    # 设置单元格的值，A1等于6(测试可知openpyxl的行和列编号从1开始计算)，B1等于7
    ws.cell(row=1, column=1).value = 'IOU'
    ws.cell("B1").value = 'Status'
    ws.cell("C1").value = 'NUM'
    # 从第1列第2行开始，写入IOU区间值
    for row in range(1, len(thresholds)+1):
        ws.cell(row=row * 2, column=1).value = (thresholds[row-1])
        ws.cell(row=row * 2 + 1, column=1).value = (thresholds[row-1])
    # 从第2列第2行开始，写入Status值
    for row in range(1, len(thresholds)+1):
        ws.cell(row=row * 2, column=2).value = 'Neg'
        ws.cell(row=row * 2 + 1, column=2).value = "Pos"
    # 从第3列第2行开始，写入NUM数据
    for row in range(1, len(thresholds)+1):
        ws.cell(row=row * 2, column=3).value = all_change_group[row-1]['Neg']
        ws.cell(row=row * 2 + 1, column=3).value = all_change_group[row-1]['Pos']
    wb.save(filename=data_xlsx)

"""
@function:将confidence变化以图形形式展现
@param param1: confidence变化文件
"""
def show_bar(data_mat):
    df = pd.read_excel(data_mat, 'Sheet')
    var = df.groupby(['IOU', 'Status']).NUM.sum()
    var.unstack().plot(kind='bar', stacked=True, color=['red', 'blue'])
    plt.title('Pos vs Neg')
    plt.show()

#matplotlib.rcParams['figure.figsize'] = (6, 8)  # 设定显示大小
plt.rcParams['image.interpolation'] = 'nearest'
plt.rcParams['image.cmap'] = 'gray'
ROOTDIR = "\\\\192.168.1.186/PedestrianData/" # 样本根目录
min_threshold = 0.1
thresholds = np.linspace( min_threshold, 1, 10 ) # IOU 区间段
all_change_group = []  # 初始化
for i in range(0, len(thresholds), 1):
    all_change_group.append({'Neg': 0, 'Pos': 0})
conf_threshold = 0.5 # 分类置信度阈值

if __name__ == "__main__":
    save_data("../Data_0810/IOU_ALL_image_List.txt",
              "../Data_0810/result_ALL_image_List.txt",
              "../Data_0810/one_confidence_change_with_IOU_statistic.xlsx")
    show_bar("../Data_0810/one_confidence_change_with_IOU_statistic.xlsx")