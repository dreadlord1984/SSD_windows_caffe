# -*- coding: utf-8 -*-
import numpy as np
import pylab as plt
from openpyxl import Workbook
import prettyplotlib as ppl
import xml.etree.cElementTree as et
import scipy.io
import pandas as pd

"""
@function:从xml文件中读取box信息
@param param1: xml文件
@return: boxes
"""
def readXML(xml_name):
    tree = et.parse(xml_name) #打开xml文档
    # 得到文档元素对象
    root = tree.getroot()
    size = root.find('size')  # 找到root节点下的size节点
    width = int(size.find('width').text)  # 子节点下节点width的值
    height = int(size.find('height').text)  # 子节点下节点height的值

    boundingBox = []
    for object in root.findall('object'):  # 找到root节点下的所有object节点
        bndbox = object.find('bndbox')  # 子节点下属性bndbox的值
        xmin = bndbox.find('xmin').text
        ymin = bndbox.find('ymin').text
        xmax = bndbox.find('xmax').text
        ymax = bndbox.find('ymax').text
        boundingBox.append([int(xmin), int(ymin), int(xmax), int(ymax)])
    return boundingBox, width, height

"""
@function:将匹配合并结果和检测结果比较分类confidence变化
@param param1: 匹配合并列表文件
@param param2: 检测结果列表文件
@param param3: 待保存gt boxes和prior boxes分布
@param param4: 待保存confidence变化文件
"""
def save_data(priorList, resultList, data_mat, data_xlsx):
    with open(priorList) as fp1, open(resultList) as fp2:  # 对于每个测试图片
        for priorFile, resultFile in zip(fp1, fp2):  # 每一行匹配数据 resultFile
            prior_datas = priorFile.strip().split('\t')
            result_datas = resultFile.strip().split('\t')
            # img_name = ROOTDIR + prior_datas[0]
            # print img_name.decode("gb2312")
            xml_name = ROOTDIR + prior_datas[1]
            # 1. 统计各个gt box面积占比区间box数量
            true_boxes, width, height = readXML(xml_name)  # 所有的ground truth boxes
            gt_boxes = []  # 样本所有gt box坐标 [[ratio_index, [xmin, ymin, xmax, ymax]]...]
            # total_gt_num += len(true_boxes)
            for gt_box in true_boxes:
                area_ratio = float((gt_box[2] - gt_box[0]) * (gt_box[3] - gt_box[1])) / (width * height)
                for i in range(0, len(area_thresholds), 1):  # 判断ratio区间段
                    if (area_ratio <= area_thresholds[i]):
                        area_gt_num[i] += 1
                        gt_boxes.append([i, gt_box])
                        break
            # 2.统计各个gt box面积占比区间IOU>0.5的prior box数量
            prior_boxes_total = int(prior_datas[2])  # 匹配box数量
            # total_prior_num += prior_boxes_total
            for i in range(0, prior_boxes_total, 1):
                gt_box_index = int(prior_datas[7 * i + 9])  # 当前匹配的gt box序号（从0开始）
                IOU = float(prior_datas[7 * i + 4])
###############################################################################################
                if IOU >= 0.1:
###############################################################################################
                    area_prior_num[gt_boxes[gt_box_index][0]] += 1
                    conf = float(result_datas[6 * i + 2])  # 分类置信度
                    for k in range(0, len(conf_thresholds), 1):  # 对于每个分类置信阈值
                        if conf >= conf_thresholds[k]:
                            all_change_group[k][gt_boxes[gt_box_index][0]]['Pos'] += 1
                        else:
                            all_change_group[k][gt_boxes[gt_box_index][0]]['Neg']  += 1
    # 3.将数据保存
    scipy.io.savemat(data_mat, {'gt_prior_statistic': gt_prior_statistic})

    # 在内存中创建一个workbook对象，而且会至少创建一个 worksheet
    wb = Workbook()
    for k in range(0, len(conf_thresholds), 1):
        # 获取当前活跃的worksheet,默认就是第一个worksheet
        ws = wb.create_sheet()
        # 设置单元格的值，A1等于6(测试可知openpyxl的行和列编号从1开始计算)，B1等于7
        ws.cell(row=1, column=1).value = 'IOU'
        ws.cell("B1").value = 'Status'
        ws.cell("C1").value = 'NUM'
        # 从第1列第2行开始，写入IOU区间值
        for row in range(1, len(area_thresholds) + 1):
            ws.cell(row=row * 2, column=1).value = (area_thresholds[row - 1])
            ws.cell(row=row * 2 + 1, column=1).value = (area_thresholds[row - 1])
        # 从第2列第2行开始，写入Status值
        for row in range(1, len(area_thresholds) + 1):
            ws.cell(row=row * 2, column=2).value = 'Neg'
            ws.cell(row=row * 2 + 1, column=2).value = "Pos"
        # 从第3列第2行开始，写入NUM数据
        for row in range(1, len(area_thresholds) + 1):
            ws.cell(row=row * 2, column=3).value = all_change_group[k][row - 1]['Neg']
            ws.cell(row=row * 2 + 1, column=3).value = all_change_group[k][row - 1]['Pos']
    wb.save(filename=data_xlsx)

"""
@function:
1.将gt boxes和prior boxes分布绘制
2.将各个conf阈值条件下confidence的变化以recall曲线展现
@param param1: gt boxes和prior boxes分布
@param param2: confidence变化文件
"""
def draw_curve(data_mat, data_xlsx):
    # 1. 绘制gt boxes和prior boxes分布
    gt_prior_statistic = scipy.io.loadmat(data_mat)
    gt_prior_statistic = gt_prior_statistic['gt_prior_statistic']
    total_gt_num = 0
    total_prior_num = 0
    for num in gt_prior_statistic[0]:
        total_gt_num += num
    for num in gt_prior_statistic[1]:
        total_prior_num += num
    fig, axes = plt.subplots(nrows=2, figsize=(8, 12))
    strand_names = ['ground truth distribution:%d' % total_gt_num, 'prior box distribution:%d' % total_prior_num]
    yalbel_names = ['gt box num', '$IOU>0.1$ prioe box num']
    labels = [area_thresholds[i] for i in s_ids]
    for ax, strand_name, ylabel_name, statistic in zip(axes, strand_names, yalbel_names, gt_prior_statistic):
        ppl.bar(ax, s_ids, statistic,
                annotate=True,
                grid='y', xticklabels=labels,
                color=[colors[i] for i in s_ids])
        ax.set_title(strand_name)
        ax.set_ylabel(ylabel_name)
        ax.set_xlabel('gt box area ratio')

    savename1 = data_mat[:data_mat.rfind("\\")] + "\\prior_gt_statistic.png"
    plt.savefig(savename1)

    # 2. 绘制不同confidence阈值下的recall曲线
    fig1, ax1 = plt.subplots(nrows=1, figsize=(10, 12))
    prior_num = []
    for k in range(0, len(conf_thresholds), 1):
        recalls = []
        sheet_name = 'Sheet'+ str(k+1)
        df = pd.read_excel(data_xlsx, sheet_name)
        var = list(df['NUM'])
        for j in range(0, len(area_thresholds), 1):
            if k==0:
                prior_num.append(var[2*j+1] + var[2*j])
            recall = float(var[2*j+1]) / (float(var[2*j+1]) + float(var[2*j]))
            recalls.append(recall)
        ax1.plot(s_ids, recalls, lw=2, color=colors[k], label=str(conf_thresholds[k])) # 绘制每一条recall曲线
        plt.annotate(conf_thresholds[k], xy=(s_ids[len(s_ids)/2], recalls[len(recalls)/2]),
                     xytext=(s_ids[len(s_ids)/2], recalls[len(recalls)/2]),
                     arrowprops = dict(facecolor="r", headlength=5, headwidth=5, width=2))

    plt.grid()
    plt.xticks(s_ids, area_thresholds)
    plt.xlabel('gt box area ratio')
    plt.ylabel('Recall')
    # plt.title('area-Recall')
    plt.legend(loc="upper left")
    ax2 = ax1.twiny()
    # ax2.set_xlabel("prior boxes num")
    plt.xticks(s_ids, prior_num, rotation=10)
    savename2 =  data_mat[:data_mat.rfind("\\")] + "\\ALL_area_recall.png"
    plt.savefig(savename2)
    plt.show()


ROOTDIR = "\\\\192.168.1.186/PedestrianData/" # 样本所在根目录
area_thresholds = np.array([0.0025, 0.005, 0.01, 0.015, 0.02, 0.04, 0.08, 0.1, 0.25, 1.0],dtype=np.float64) # area 区间
s_ids = np.arange(len(area_thresholds))
area_gt_num = np.zeros(len(area_thresholds),dtype=np.int32)
area_prior_num = np.zeros(len(area_thresholds),dtype=np.int32)
gt_prior_statistic = []
gt_prior_statistic.append(area_gt_num)
gt_prior_statistic.append(area_prior_num)
#colorbar =(0.4, 0.7607843137254902, 0.6470588235294118)
colors = plt.cm.hsv(np.linspace(0, 1, 10)).tolist()
min_conf_threshold = 0.1 # 最小分类置信度阈值
conf_thresholds = np.linspace( min_conf_threshold, 1, 10 ) # 分类置信度阈值
all_change_group =  [[] for x in range(len(conf_thresholds))]  # 初始化
for j in range(0, len(conf_thresholds), 1):
    for i in range(0, len(area_thresholds), 1):
        all_change_group[j].append({'Neg': 0, 'Pos': 0})


if __name__ == "__main__":
    save_data("..\\View\\COMPARE2\\add_prior_gamma2_D_new_P5N4D15E4\\IOU_ALL_image_List.txt",
              "..\\View\\COMPARE2\\add_prior_gamma2_D_new_P5N4D15E4\\result_ALL_image_List.txt",
              "..\\View\\COMPARE2\\add_prior_gamma2_D_new_P5N4D15E4\\object_confidence_change_curve.mat",
              "..\\View\\COMPARE2\\add_prior_gamma2_D_new_P5N4D15E4\\object_confidence_change_curve.xlsx")
    draw_curve("..\\View\\COMPARE2\\add_prior_gamma2_D_new_P5N4D15E4\\object_confidence_change_curve.mat",
               "..\\View\\COMPARE2\\add_prior_gamma2_D_new_P5N4D15E4\\object_confidence_change_curve.xlsx")